import csv
import io
import json
from datetime import datetime, date, timedelta
from typing import Optional

from sqlalchemy import func, asc, desc, cast, Date, text
from sqlalchemy.orm import Session

from app.modules.hr.models import (
    AttendanceRecord, AttendanceRegularization, AttendancePolicy,
    Shift, ShiftRoster, GeofenceLocation, OvertimeRequest,
    AttendanceException, Holiday, WeekendConfig, BiometricDevice,
    AttendanceAuditLog, Employee, Department,
    AttendanceStatus, CorrectionType, RegularizationStatus,
    ShiftType, OvertimeStatus, ExceptionStatus, ExceptionType,
)
from app.modules.hr.schemas import (
    AttendanceCreate, AttendanceUpdate,
    RegularizationCreate,
    AttendancePolicyCreate, AttendancePolicyUpdate,
    ShiftCreate, ShiftUpdate,
    ShiftRosterCreate,
    GeofenceCreate, GeofenceUpdate,
    OvertimeCreate,
    HolidayCreate, HolidayUpdate,
    WeekendConfigCreate,
    BiometricDeviceCreate, BiometricDeviceUpdate,
    SuccessResponse,
)
from app.core.exceptions import NotFoundException, AlreadyExistsException, BadRequestException
from app.core.sanitize import sanitize_input, sanitize_dict


# ── DASHBOARD ─────────────────────────────────────────────────────────────────

def get_attendance_dashboard(db: Session) -> dict:
    today = date.today()

    present_today = db.query(func.count(AttendanceRecord.id)).filter(
        AttendanceRecord.date == today,
        AttendanceRecord.status == AttendanceStatus.PRESENT,
    ).scalar() or 0

    absent_today = db.query(func.count(AttendanceRecord.id)).filter(
        AttendanceRecord.date == today,
        AttendanceRecord.status == AttendanceStatus.ABSENT,
    ).scalar() or 0

    on_leave_count = db.query(func.count(AttendanceRecord.id)).filter(
        AttendanceRecord.date == today,
        AttendanceRecord.status == AttendanceStatus.ON_LEAVE,
    ).scalar() or 0

    remote_count = db.query(func.count(AttendanceRecord.id)).filter(
        AttendanceRecord.date == today,
        AttendanceRecord.status == AttendanceStatus.REMOTE,
    ).scalar() or 0

    total_emp = db.query(func.count(Employee.id)).filter(Employee.is_active == True).scalar() or 1
    attendance_percentage = round((present_today / total_emp) * 100, 2) if total_emp else 0.0

    late_arrivals = 0
    early_departures = 0
    overtime_count = db.query(func.count(OvertimeRequest.id)).filter(
        OvertimeRequest.date == today,
        OvertimeRequest.status == OvertimeStatus.APPROVED,
    ).scalar() or 0

    avg_hours = db.query(
        func.avg(
            func.TIMESTAMPDIFF(text('SECOND'), AttendanceRecord.check_in, AttendanceRecord.check_out) / 3600
        )
    ).filter(
        AttendanceRecord.date == today,
        AttendanceRecord.check_in.isnot(None),
        AttendanceRecord.check_out.isnot(None),
    ).scalar() or 0.0
    avg_working_hours = round(float(avg_hours), 2)

    dept_breakdown = (
        db.query(
            Department.name,
            func.count(AttendanceRecord.id),
        )
        .select_from(AttendanceRecord)
        .join(Employee, AttendanceRecord.employee_id == Employee.id)
        .join(Department, Employee.department_id == Department.id)
        .filter(AttendanceRecord.date == today)
        .group_by(Department.name)
        .all()
    )

    shift_util = (
        db.query(
            Shift.name,
            func.count(AttendanceRecord.id),
        )
        .select_from(Shift)
        .outerjoin(ShiftRoster, Shift.id == ShiftRoster.shift_id)
        .outerjoin(AttendanceRecord, (AttendanceRecord.employee_id == ShiftRoster.employee_id) & (AttendanceRecord.date == today))
        .group_by(Shift.name)
        .all()
    )

    return {
        "present_today": present_today,
        "absent_today": absent_today,
        "late_arrivals": late_arrivals,
        "early_departures": early_departures,
        "on_leave_count": on_leave_count,
        "remote_count": remote_count,
        "overtime_count": overtime_count,
        "attendance_percentage": attendance_percentage,
        "avg_working_hours": avg_working_hours,
        "department_breakdown": [{"department": d, "count": c} for d, c in dept_breakdown],
        "shift_utilization": [{"shift": s, "count": c} for s, c in shift_util],
    }


# ── ATTENDANCE RECORDS CRUD ──────────────────────────────────────────────────

SORTABLE_FIELDS_RECORDS = {
    "id": AttendanceRecord.id,
    "employee_id": AttendanceRecord.employee_id,
    "date": AttendanceRecord.date,
    "status": AttendanceRecord.status,
    "check_in": AttendanceRecord.check_in,
    "check_out": AttendanceRecord.check_out,
    "created_at": AttendanceRecord.created_at,
}


def _get_records_query(
    db: Session, search=None, status=None, department=None,
    date_from=None, date_to=None, employee_id=None,
):
    query = db.query(AttendanceRecord)

    if search:
        stmt = f"%{search}%"
        query = query.join(Employee, AttendanceRecord.employee_id == Employee.id).filter(
            (Employee.first_name.ilike(stmt)) |
            (Employee.last_name.ilike(stmt)) |
            (Employee.employee_code.ilike(stmt))
        )

    if status:
        query = query.filter(AttendanceRecord.status == status)

    if department:
        query = query.join(Employee, AttendanceRecord.employee_id == Employee.id).filter(
            Employee.department_id == Department.id,
            Department.name == department,
        )

    if date_from:
        query = query.filter(AttendanceRecord.date >= date_from)

    if date_to:
        query = query.filter(AttendanceRecord.date <= date_to)

    if employee_id:
        query = query.filter(AttendanceRecord.employee_id == employee_id)

    return query


def get_attendance_records(
    db: Session,
    page: int = 1,
    per_page: int = 20,
    search: Optional[str] = None,
    status: Optional[AttendanceStatus] = None,
    department: Optional[str] = None,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    employee_id: Optional[int] = None,
    sort_by: Optional[str] = "date",
    sort_order: Optional[str] = "desc",
) -> dict:
    per_page = min(per_page, 100)
    query = _get_records_query(db, search, status, department, date_from, date_to, employee_id)
    total = query.count()

    sort_col = SORTABLE_FIELDS_RECORDS.get(sort_by, AttendanceRecord.date)
    sort_fn = desc if sort_order == "desc" else asc
    records = query.order_by(sort_fn(sort_col)).offset((page - 1) * per_page).limit(per_page).all()

    items = []
    for r in records:
        items.append({
            "id": r.id,
            "employee_id": r.employee_id,
            "date": r.date,
            "status": r.status,
            "check_in": r.check_in,
            "check_out": r.check_out,
            "notes": r.notes,
            "created_at": r.created_at,
            "employee_name": r.employee.full_name if r.employee else None,
        })

    return {"total": total, "page": page, "per_page": per_page, "items": items}


def create_attendance_record(db: Session, data: AttendanceCreate, created_by: int = None) -> AttendanceRecord:
    raw = data.model_dump()
    safe = sanitize_dict(raw)
    record = AttendanceRecord(**safe)
    db.add(record)
    db.commit()
    db.refresh(record)
    _log_audit(db, "CREATE", "AttendanceRecord", record.id, record.employee_id, raw, created_by)
    return record


def get_attendance_record_by_id(db: Session, record_id: int) -> AttendanceRecord:
    record = db.query(AttendanceRecord).filter(AttendanceRecord.id == record_id).first()
    if not record:
        raise NotFoundException("AttendanceRecord", record_id)
    return record


def update_attendance_record(db: Session, record_id: int, data: AttendanceUpdate, updated_by: int = None) -> AttendanceRecord:
    record = get_attendance_record_by_id(db, record_id)
    update_data = sanitize_dict(data.model_dump(exclude_unset=True))
    old_vals = {"status": record.status.value if record.status else None, "notes": record.notes}
    for field, value in update_data.items():
        setattr(record, field, value)
    db.commit()
    db.refresh(record)
    _log_audit(db, "UPDATE", "AttendanceRecord", record_id, record.employee_id,
               {"old": old_vals, "new": update_data}, updated_by)
    return record


def delete_attendance_record(db: Session, record_id: int, deleted_by: int = None) -> None:
    record = get_attendance_record_by_id(db, record_id)
    record.deleted_at = datetime.utcnow() if hasattr(record, 'deleted_at') else None
    db.commit()
    _log_audit(db, "DELETE", "AttendanceRecord", record_id, record.employee_id, None, deleted_by)


# ── CLOCK IN/OUT ─────────────────────────────────────────────────────────────

def clock_in(db: Session, employee_id: int, clock_in_time: Optional[datetime] = None, created_by: int = None) -> AttendanceRecord:
    today = date.today()
    existing = db.query(AttendanceRecord).filter(
        AttendanceRecord.employee_id == employee_id,
        AttendanceRecord.date == today,
        AttendanceRecord.check_in.isnot(None),
    ).first()
    if existing:
        raise BadRequestException("Employee already clocked in today")
    now = clock_in_time or datetime.utcnow()
    record = AttendanceRecord(
        employee_id=employee_id,
        date=today,
        status=AttendanceStatus.PRESENT,
        check_in=now,
        created_by=created_by,
    )
    db.add(record)
    db.commit()
    db.refresh(record)
    return record


def clock_out(db: Session, record_id: int, clock_out_time: Optional[datetime] = None) -> AttendanceRecord:
    record = get_attendance_record_by_id(db, record_id)
    if record.check_out:
        raise BadRequestException("Already clocked out for this record")
    record.check_out = clock_out_time or datetime.utcnow()
    db.commit()
    db.refresh(record)
    return record


def break_start(db: Session, record_id: int, break_start_time: Optional[datetime] = None) -> AttendanceRecord:
    record = get_attendance_record_by_id(db, record_id)
    if not record.check_in:
        raise BadRequestException("Must clock in first")
    record.break_start = break_start_time or datetime.utcnow()
    db.commit()
    db.refresh(record)
    return record


def break_end(db: Session, record_id: int, break_end_time: Optional[datetime] = None) -> AttendanceRecord:
    record = get_attendance_record_by_id(db, record_id)
    if not hasattr(record, 'break_start') or not record.break_start:
        raise BadRequestException("No break started for this record")
    record.break_end = break_end_time or datetime.utcnow()
    db.commit()
    db.refresh(record)
    return record


# ── ATTENDANCE REGULARIZATION ────────────────────────────────────────────────

def get_regularizations(
    db: Session, page: int = 1, per_page: int = 20,
    status: Optional[RegularizationStatus] = None,
    employee_id: Optional[int] = None,
    correction_type: Optional[CorrectionType] = None,
) -> dict:
    per_page = min(per_page, 100)
    query = db.query(AttendanceRegularization)
    if status:
        query = query.filter(AttendanceRegularization.status == status)
    if employee_id:
        query = query.filter(AttendanceRegularization.employee_id == employee_id)
    if correction_type:
        query = query.filter(AttendanceRegularization.correction_type == correction_type)
    total = query.count()
    items = query.order_by(AttendanceRegularization.created_at.desc()).offset((page - 1) * per_page).limit(per_page).all()
    result = []
    for r in items:
        result.append({
            "id": r.id,
            "employee_id": r.employee_id,
            "attendance_record_id": r.attendance_record_id,
            "correction_type": r.correction_type,
            "date": r.date,
            "expected_check_in": r.expected_check_in,
            "expected_check_out": r.expected_check_out,
            "actual_check_in": r.actual_check_in,
            "actual_check_out": r.actual_check_out,
            "reason": r.reason,
            "status": r.status,
            "manager_id": r.manager_id,
            "manager_approved_at": r.manager_approved_at,
            "hr_approved_at": r.hr_approved_at,
            "rejected_by": r.rejected_by,
            "rejected_at": r.rejected_at,
            "rejection_reason": r.rejection_reason,
            "created_at": r.created_at,
            "updated_at": r.updated_at,
            "employee_name": r.employee.full_name if r.employee else None,
        })
    return {"total": total, "page": page, "per_page": per_page, "items": result}


def create_regularization(db: Session, data: RegularizationCreate, created_by: int = None) -> AttendanceRegularization:
    reg_data = data.model_dump()
    if reg_data.get("employee_id") is None:
        reg_data["employee_id"] = created_by
    req = AttendanceRegularization(**reg_data)
    db.add(req)
    db.commit()
    db.refresh(req)
    _log_audit(db, "CREATE", "AttendanceRegularization", req.id, req.employee_id, reg_data, created_by)
    return req


def get_regularization_by_id(db: Session, reg_id: int) -> AttendanceRegularization:
    req = db.query(AttendanceRegularization).filter(AttendanceRegularization.id == reg_id).first()
    if not req:
        raise NotFoundException("AttendanceRegularization", reg_id)
    return req


def approve_regularization_manager(db: Session, reg_id: int, manager_id: int) -> AttendanceRegularization:
    req = get_regularization_by_id(db, reg_id)
    if req.status != RegularizationStatus.PENDING:
        raise BadRequestException("Regularization is not in PENDING status")
    req.status = RegularizationStatus.MANAGER_APPROVED
    req.manager_id = manager_id
    req.manager_approved_at = datetime.utcnow()
    db.commit()
    db.refresh(req)
    return req


def approve_regularization_hr(db: Session, reg_id: int, hr_id: int) -> AttendanceRegularization:
    req = get_regularization_by_id(db, reg_id)
    if req.status not in (RegularizationStatus.PENDING, RegularizationStatus.MANAGER_APPROVED):
        raise BadRequestException("Regularization cannot be HR-approved in its current status")
    req.status = RegularizationStatus.APPROVED
    req.hr_approved_at = datetime.utcnow()
    db.commit()
    db.refresh(req)
    if req.attendance_record_id:
        att = db.query(AttendanceRecord).filter(AttendanceRecord.id == req.attendance_record_id).first()
        if att:
            att.status = AttendanceStatus.PRESENT
            if req.expected_check_in:
                att.check_in = req.expected_check_in
            if req.expected_check_out:
                att.check_out = req.expected_check_out
            db.commit()
    return req


def reject_regularization(db: Session, reg_id: int, rejected_by: int, rejection_reason: Optional[str] = None) -> AttendanceRegularization:
    req = get_regularization_by_id(db, reg_id)
    if req.status in (RegularizationStatus.APPROVED, RegularizationStatus.REJECTED, RegularizationStatus.CANCELLED):
        raise BadRequestException("Regularization cannot be rejected in its current status")
    req.status = RegularizationStatus.REJECTED
    req.rejected_by = rejected_by
    req.rejected_at = datetime.utcnow()
    req.rejection_reason = rejection_reason
    db.commit()
    db.refresh(req)
    return req


def cancel_regularization(db: Session, reg_id: int) -> AttendanceRegularization:
    req = get_regularization_by_id(db, reg_id)
    if req.status in (RegularizationStatus.APPROVED, RegularizationStatus.REJECTED, RegularizationStatus.CANCELLED):
        raise BadRequestException("Regularization cannot be cancelled in its current status")
    req.status = RegularizationStatus.CANCELLED
    db.commit()
    db.refresh(req)
    return req


# ── ATTENDANCE POLICIES ──────────────────────────────────────────────────────

def get_attendance_policies(db: Session) -> list[AttendancePolicy]:
    return db.query(AttendancePolicy).order_by(AttendancePolicy.name).all()


def create_attendance_policy(db: Session, data: AttendancePolicyCreate, created_by: int = None) -> AttendancePolicy:
    existing = db.query(AttendancePolicy).filter(AttendancePolicy.name.ilike(data.name)).first()
    if existing:
        raise AlreadyExistsException("AttendancePolicy", "name")
    policy = AttendancePolicy(**data.model_dump(), created_by=created_by)
    db.add(policy)
    db.commit()
    db.refresh(policy)
    return policy


def get_attendance_policy_by_id(db: Session, policy_id: int) -> AttendancePolicy:
    policy = db.query(AttendancePolicy).filter(AttendancePolicy.id == policy_id).first()
    if not policy:
        raise NotFoundException("AttendancePolicy", policy_id)
    return policy


def update_attendance_policy(db: Session, policy_id: int, data: AttendancePolicyUpdate, updated_by: int = None) -> AttendancePolicy:
    policy = get_attendance_policy_by_id(db, policy_id)
    update_data = sanitize_dict(data.model_dump(exclude_unset=True))
    for field, value in update_data.items():
        setattr(policy, field, value)
    db.commit()
    db.refresh(policy)
    return policy


def delete_attendance_policy(db: Session, policy_id: int) -> None:
    policy = get_attendance_policy_by_id(db, policy_id)
    db.delete(policy)
    db.commit()


# ── SHIFTS ───────────────────────────────────────────────────────────────────

def get_shifts(db: Session) -> list[Shift]:
    return db.query(Shift).order_by(Shift.name).all()


def create_shift(db: Session, data: ShiftCreate, created_by: int = None) -> Shift:
    shift = Shift(**data.model_dump(), created_by=created_by)
    db.add(shift)
    db.commit()
    db.refresh(shift)
    return shift


def get_shift_by_id(db: Session, shift_id: int) -> Shift:
    shift = db.query(Shift).filter(Shift.id == shift_id).first()
    if not shift:
        raise NotFoundException("Shift", shift_id)
    return shift


def update_shift(db: Session, shift_id: int, data: ShiftUpdate, updated_by: int = None) -> Shift:
    shift = get_shift_by_id(db, shift_id)
    update_data = sanitize_dict(data.model_dump(exclude_unset=True))
    for field, value in update_data.items():
        setattr(shift, field, value)
    db.commit()
    db.refresh(shift)
    return shift


def delete_shift(db: Session, shift_id: int) -> None:
    shift = get_shift_by_id(db, shift_id)
    db.delete(shift)
    db.commit()


# ── SHIFT ROSTERS ────────────────────────────────────────────────────────────

def get_shift_rosters(
    db: Session, page: int = 1, per_page: int = 20,
    date_filter: Optional[date] = None,
    employee_id: Optional[int] = None,
    shift_id: Optional[int] = None,
) -> dict:
    per_page = min(per_page, 100)
    query = db.query(ShiftRoster)
    if date_filter:
        query = query.filter(ShiftRoster.date == date_filter)
    if employee_id:
        query = query.filter(ShiftRoster.employee_id == employee_id)
    if shift_id:
        query = query.filter(ShiftRoster.shift_id == shift_id)
    total = query.count()
    items = query.order_by(ShiftRoster.date.desc()).offset((page - 1) * per_page).limit(per_page).all()
    result = []
    for r in items:
        result.append({
            "id": r.id,
            "employee_id": r.employee_id,
            "shift_id": r.shift_id,
            "date": r.date,
            "is_active": r.is_active,
            "assigned_by": r.assigned_by,
            "created_at": r.created_at,
            "updated_at": r.updated_at,
            "employee_name": r.employee.full_name if r.employee else None,
            "shift_name": r.shift.name if r.shift else None,
        })
    return {"total": total, "page": page, "per_page": per_page, "items": result}


def create_shift_roster(db: Session, data: ShiftRosterCreate, assigned_by: int = None) -> ShiftRoster:
    roster = ShiftRoster(**data.model_dump(), assigned_by=assigned_by)
    db.add(roster)
    db.commit()
    db.refresh(roster)
    return roster


def bulk_assign_shifts(db: Session, assignments: list[dict], assigned_by: int = None) -> list[ShiftRoster]:
    created = []
    for item in assignments:
        safe = sanitize_dict(item)
        roster = ShiftRoster(
            employee_id=safe.get("employee_id"),
            shift_id=safe.get("shift_id"),
            date=safe.get("date"),
            assigned_by=assigned_by,
        )
        db.add(roster)
        created.append(roster)
    db.commit()
    for r in created:
        db.refresh(r)
    return created


def delete_shift_roster(db: Session, roster_id: int) -> None:
    roster = db.query(ShiftRoster).filter(ShiftRoster.id == roster_id).first()
    if not roster:
        raise NotFoundException("ShiftRoster", roster_id)
    db.delete(roster)
    db.commit()


# ── MY ATTENDANCE ────────────────────────────────────────────────────────────

def get_my_attendance(db: Session, employee_id: int, page: int = 1, per_page: int = 20) -> dict:
    per_page = min(per_page, 100)
    query = db.query(AttendanceRecord).filter(AttendanceRecord.employee_id == employee_id)
    total = query.count()
    records = query.order_by(AttendanceRecord.date.desc()).offset((page - 1) * per_page).limit(per_page).all()
    items = []
    for r in records:
        items.append({
            "id": r.id,
            "employee_id": r.employee_id,
            "date": r.date,
            "status": r.status,
            "check_in": r.check_in,
            "check_out": r.check_out,
            "notes": r.notes,
            "created_at": r.created_at,
        })
    return {"total": total, "page": page, "per_page": per_page, "items": items}


def get_employee_summary(db: Session, employee_id: int, month: int, year: int) -> dict:
    records = db.query(AttendanceRecord).filter(
        AttendanceRecord.employee_id == employee_id,
        func.extract('month', AttendanceRecord.date) == month,
        func.extract('year', AttendanceRecord.date) == year,
    ).all()

    present = sum(1 for r in records if r.status == AttendanceStatus.PRESENT)
    absent = sum(1 for r in records if r.status == AttendanceStatus.ABSENT)
    on_leave = sum(1 for r in records if r.status == AttendanceStatus.ON_LEAVE)
    remote = sum(1 for r in records if r.status == AttendanceStatus.REMOTE)
    total_days = len(records) or 1

    total_hours = 0.0
    for r in records:
        if r.check_in and r.check_out:
            diff = (r.check_out - r.check_in).total_seconds() / 3600
            total_hours += diff

    return {
        "employee_id": employee_id,
        "month": month,
        "year": year,
        "present": present,
        "absent": absent,
        "on_leave": on_leave,
        "remote": remote,
        "total_days": total_days,
        "attendance_percentage": round((present / total_days) * 100, 2),
        "total_hours": round(total_hours, 2),
        "avg_hours_per_day": round(total_hours / total_days, 2) if total_days else 0.0,
    }


def get_employee_history(db: Session, employee_id: int, page: int = 1, per_page: int = 20,
                         date_from: Optional[date] = None, date_to: Optional[date] = None) -> dict:
    per_page = min(per_page, 100)
    query = db.query(AttendanceRecord).filter(AttendanceRecord.employee_id == employee_id)
    if date_from:
        query = query.filter(AttendanceRecord.date >= date_from)
    if date_to:
        query = query.filter(AttendanceRecord.date <= date_to)
    total = query.count()
    records = query.order_by(AttendanceRecord.date.desc()).offset((page - 1) * per_page).limit(per_page).all()
    items = []
    for r in records:
        items.append({
            "id": r.id,
            "date": r.date,
            "status": r.status,
            "check_in": r.check_in,
            "check_out": r.check_out,
            "notes": r.notes,
        })
    return {"total": total, "page": page, "per_page": per_page, "items": items}


def get_employee_score(db: Session, employee_id: int) -> dict:
    total_records = db.query(func.count(AttendanceRecord.id)).filter(
        AttendanceRecord.employee_id == employee_id,
    ).scalar() or 0

    present = db.query(func.count(AttendanceRecord.id)).filter(
        AttendanceRecord.employee_id == employee_id,
        AttendanceRecord.status == AttendanceStatus.PRESENT,
    ).scalar() or 0

    late_count = db.query(func.count(AttendanceRecord.id)).filter(
        AttendanceRecord.employee_id == employee_id,
        AttendanceRecord.status == AttendanceStatus.PRESENT,
        AttendanceRecord.check_in.isnot(None),
    ).all()

    score = round((present / max(total_records, 1)) * 100, 2) if total_records else 0.0

    return {
        "employee_id": employee_id,
        "total_records": total_records,
        "present_count": present,
        "attendance_score": score,
        "grade": "A" if score >= 90 else "B" if score >= 75 else "C" if score >= 50 else "D",
    }


# ── BIOMETRIC DEVICES ────────────────────────────────────────────────────────

def get_biometric_devices(db: Session) -> list[BiometricDevice]:
    return db.query(BiometricDevice).order_by(BiometricDevice.name).all()


def register_biometric_device(db: Session, data: BiometricDeviceCreate, created_by: int = None) -> BiometricDevice:
    existing = db.query(BiometricDevice).filter(BiometricDevice.device_id == data.device_id).first()
    if existing:
        raise AlreadyExistsException("BiometricDevice", "device_id")
    device = BiometricDevice(**data.model_dump(), created_by=created_by)
    db.add(device)
    db.commit()
    db.refresh(device)
    return device


def update_biometric_device(db: Session, device_id: int, data: BiometricDeviceUpdate, updated_by: int = None) -> BiometricDevice:
    device = db.query(BiometricDevice).filter(BiometricDevice.id == device_id).first()
    if not device:
        raise NotFoundException("BiometricDevice", device_id)
    update_data = sanitize_dict(data.model_dump(exclude_unset=True))
    for field, value in update_data.items():
        setattr(device, field, value)
    db.commit()
    db.refresh(device)
    return device


def delete_biometric_device(db: Session, device_id: int) -> None:
    device = db.query(BiometricDevice).filter(BiometricDevice.id == device_id).first()
    if not device:
        raise NotFoundException("BiometricDevice", device_id)
    db.delete(device)
    db.commit()


def sync_biometric_logs(db: Session, device_id: int, logs: list[dict], synced_by: int = None) -> dict:
    device = db.query(BiometricDevice).filter(BiometricDevice.id == device_id).first()
    if not device:
        raise NotFoundException("BiometricDevice", device_id)
    imported = 0
    for log in logs:
        emp_id = log.get("employee_id")
        punch_time = log.get("punch_time")
        if emp_id and punch_time:
            punch_dt = datetime.fromisoformat(punch_time) if isinstance(punch_time, str) else punch_time
            punch_date = punch_dt.date()
            existing = db.query(AttendanceRecord).filter(
                AttendanceRecord.employee_id == emp_id,
                AttendanceRecord.date == punch_date,
            ).first()
            if existing:
                if not existing.check_in:
                    existing.check_in = punch_dt
                else:
                    existing.check_out = punch_dt
            else:
                record = AttendanceRecord(
                    employee_id=emp_id,
                    date=punch_date,
                    status=AttendanceStatus.PRESENT,
                    check_in=punch_dt,
                )
                db.add(record)
            imported += 1
    device.last_sync = datetime.utcnow()
    db.commit()
    return {"imported": imported, "device_id": device_id}


def import_biometric_logs(db: Session, logs: list[dict], imported_by: int = None) -> dict:
    imported = 0
    for log in logs:
        emp_id = log.get("employee_id")
        punch_time = log.get("punch_time")
        if emp_id and punch_time:
            punch_dt = datetime.fromisoformat(punch_time) if isinstance(punch_time, str) else punch_time
            punch_date = punch_dt.date()
            existing = db.query(AttendanceRecord).filter(
                AttendanceRecord.employee_id == emp_id,
                AttendanceRecord.date == punch_date,
            ).first()
            if existing:
                if not existing.check_in:
                    existing.check_in = punch_dt
                else:
                    existing.check_out = punch_dt
            else:
                record = AttendanceRecord(
                    employee_id=emp_id,
                    date=punch_date,
                    status=AttendanceStatus.PRESENT,
                    check_in=punch_dt,
                )
                db.add(record)
            imported += 1
    db.commit()
    return {"imported": imported}


def get_biometric_device_health(db: Session, device_id: int) -> dict:
    device = db.query(BiometricDevice).filter(BiometricDevice.id == device_id).first()
    if not device:
        raise NotFoundException("BiometricDevice", device_id)
    healthy = device.is_active and device.last_sync is not None
    return {
        "id": device.id,
        "name": device.name,
        "is_active": device.is_active,
        "last_sync": device.last_sync,
        "health_status": "healthy" if healthy else "unhealthy",
    }


# ── GEOFENCING ───────────────────────────────────────────────────────────────

def get_geofence_locations(db: Session) -> list[GeofenceLocation]:
    return db.query(GeofenceLocation).order_by(GeofenceLocation.name).all()


def create_geofence_location(db: Session, data: GeofenceCreate, created_by: int = None) -> GeofenceLocation:
    loc = GeofenceLocation(**data.model_dump(), created_by=created_by)
    db.add(loc)
    db.commit()
    db.refresh(loc)
    return loc


def get_geofence_location_by_id(db: Session, location_id: int) -> GeofenceLocation:
    loc = db.query(GeofenceLocation).filter(GeofenceLocation.id == location_id).first()
    if not loc:
        raise NotFoundException("GeofenceLocation", location_id)
    return loc


def update_geofence_location(db: Session, location_id: int, data: GeofenceUpdate, updated_by: int = None) -> GeofenceLocation:
    loc = get_geofence_location_by_id(db, location_id)
    update_data = sanitize_dict(data.model_dump(exclude_unset=True))
    for field, value in update_data.items():
        setattr(loc, field, value)
    db.commit()
    db.refresh(loc)
    return loc


def delete_geofence_location(db: Session, location_id: int) -> None:
    loc = get_geofence_location_by_id(db, location_id)
    db.delete(loc)
    db.commit()


# ── OVERTIME ─────────────────────────────────────────────────────────────────

def get_overtime_requests(
    db: Session, page: int = 1, per_page: int = 20,
    status: Optional[OvertimeStatus] = None,
    employee_id: Optional[int] = None,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
) -> dict:
    per_page = min(per_page, 100)
    query = db.query(OvertimeRequest)
    if status:
        query = query.filter(OvertimeRequest.status == status)
    if employee_id:
        query = query.filter(OvertimeRequest.employee_id == employee_id)
    if date_from:
        query = query.filter(OvertimeRequest.date >= date_from)
    if date_to:
        query = query.filter(OvertimeRequest.date <= date_to)
    total = query.count()
    items = query.order_by(OvertimeRequest.created_at.desc()).offset((page - 1) * per_page).limit(per_page).all()
    result = []
    for r in items:
        result.append({
            "id": r.id,
            "employee_id": r.employee_id,
            "date": r.date,
            "hours_requested": r.hours_requested,
            "hours_approved": r.hours_approved,
            "reason": r.reason,
            "status": r.status,
            "approved_by": r.approved_by,
            "approved_at": r.approved_at,
            "rejected_by": r.rejected_by,
            "rejected_at": r.rejected_at,
            "rejection_reason": r.rejection_reason,
            "created_at": r.created_at,
            "updated_at": r.updated_at,
            "employee_name": r.employee.full_name if r.employee else None,
        })
    return {"total": total, "page": page, "per_page": per_page, "items": result}


def create_overtime_request(db: Session, data: OvertimeCreate, created_by: int = None) -> OvertimeRequest:
    req = OvertimeRequest(**data.model_dump())
    db.add(req)
    db.commit()
    db.refresh(req)
    return req


def get_overtime_request_by_id(db: Session, ot_id: int) -> OvertimeRequest:
    req = db.query(OvertimeRequest).filter(OvertimeRequest.id == ot_id).first()
    if not req:
        raise NotFoundException("OvertimeRequest", ot_id)
    return req


def approve_overtime_request(db: Session, ot_id: int, approved_by: int, hours_approved: Optional[float] = None) -> OvertimeRequest:
    req = get_overtime_request_by_id(db, ot_id)
    if req.status != OvertimeStatus.PENDING:
        raise BadRequestException("Overtime request is not in PENDING status")
    req.status = OvertimeStatus.APPROVED
    req.approved_by = approved_by
    req.approved_at = datetime.utcnow()
    if hours_approved:
        req.hours_approved = hours_approved
    db.commit()
    db.refresh(req)
    return req


def reject_overtime_request(db: Session, ot_id: int, rejected_by: int, rejection_reason: Optional[str] = None) -> OvertimeRequest:
    req = get_overtime_request_by_id(db, ot_id)
    if req.status != OvertimeStatus.PENDING:
        raise BadRequestException("Overtime request is not in PENDING status")
    req.status = OvertimeStatus.REJECTED
    req.rejected_by = rejected_by
    req.rejected_at = datetime.utcnow()
    req.rejection_reason = rejection_reason
    db.commit()
    db.refresh(req)
    return req


def get_overtime_reports(db: Session, date_from: Optional[date] = None, date_to: Optional[date] = None) -> dict:
    query = db.query(OvertimeRequest)
    if date_from:
        query = query.filter(OvertimeRequest.date >= date_from)
    if date_to:
        query = query.filter(OvertimeRequest.date <= date_to)

    total_requests = query.count()
    total_approved = query.filter(OvertimeRequest.status == OvertimeStatus.APPROVED).count()
    total_hours = db.query(func.sum(OvertimeRequest.hours_requested)).filter(
        OvertimeRequest.status == OvertimeStatus.APPROVED,
    ).scalar() or 0

    by_employee = (
        db.query(
            Employee.id,
            Employee.first_name,
            Employee.last_name,
            func.sum(OvertimeRequest.hours_requested),
            func.count(OvertimeRequest.id),
        )
        .join(OvertimeRequest, Employee.id == OvertimeRequest.employee_id)
        .filter(OvertimeRequest.status == OvertimeStatus.APPROVED)
        .group_by(Employee.id)
        .all()
    )

    return {
        "total_requests": total_requests,
        "total_approved": total_approved,
        "total_hours": float(total_hours),
        "employee_breakdown": [
            {
                "employee_id": e[0],
                "employee_name": f"{e[1]} {e[2]}",
                "total_hours": float(e[3]) if e[3] else 0,
                "total_requests": e[4],
            }
            for e in by_employee
        ],
    }


# ── ATTENDANCE EXCEPTIONS ────────────────────────────────────────────────────

def get_attendance_exceptions(
    db: Session, page: int = 1, per_page: int = 20,
    status: Optional[ExceptionStatus] = None,
    exception_type: Optional[ExceptionType] = None,
    employee_id: Optional[int] = None,
) -> dict:
    per_page = min(per_page, 100)
    query = db.query(AttendanceException)
    if status:
        query = query.filter(AttendanceException.status == status)
    if exception_type:
        query = query.filter(AttendanceException.exception_type == exception_type)
    if employee_id:
        query = query.filter(AttendanceException.employee_id == employee_id)
    total = query.count()
    items = query.order_by(AttendanceException.created_at.desc()).offset((page - 1) * per_page).limit(per_page).all()
    result = []
    for r in items:
        result.append({
            "id": r.id,
            "employee_id": r.employee_id,
            "attendance_record_id": r.attendance_record_id,
            "exception_type": r.exception_type,
            "description": r.description,
            "status": r.status,
            "resolved_by": r.resolved_by,
            "resolved_at": r.resolved_at,
            "resolution_notes": r.resolution_notes,
            "escalated_to": r.escalated_to,
            "escalated_at": r.escalated_at,
            "created_at": r.created_at,
            "employee_name": r.employee.full_name if r.employee else None,
        })
    return {"total": total, "page": page, "per_page": per_page, "items": result}


def create_attendance_exception(db: Session, employee_id: int, attendance_record_id: Optional[int] = None,
                                 exception_type: ExceptionType = ExceptionType.ATTENDANCE_VIOLATION,
                                 description: Optional[str] = None) -> AttendanceException:
    exc = AttendanceException(
        employee_id=employee_id,
        attendance_record_id=attendance_record_id,
        exception_type=exception_type,
        description=description,
        status=ExceptionStatus.OPEN,
    )
    db.add(exc)
    db.commit()
    db.refresh(exc)
    return exc


def resolve_attendance_exception(db: Session, exc_id: int, resolved_by: int, resolution_notes: Optional[str] = None) -> AttendanceException:
    exc = db.query(AttendanceException).filter(AttendanceException.id == exc_id).first()
    if not exc:
        raise NotFoundException("AttendanceException", exc_id)
    exc.status = ExceptionStatus.RESOLVED
    exc.resolved_by = resolved_by
    exc.resolved_at = datetime.utcnow()
    exc.resolution_notes = resolution_notes
    db.commit()
    db.refresh(exc)
    return exc


def escalate_attendance_exception(db: Session, exc_id: int, escalated_to: int) -> AttendanceException:
    exc = db.query(AttendanceException).filter(AttendanceException.id == exc_id).first()
    if not exc:
        raise NotFoundException("AttendanceException", exc_id)
    exc.status = ExceptionStatus.ESCALATED
    exc.escalated_to = escalated_to
    exc.escalated_at = datetime.utcnow()
    db.commit()
    db.refresh(exc)
    return exc


def get_attendance_exception_by_id(db: Session, exc_id: int) -> AttendanceException:
    exc = db.query(AttendanceException).filter(AttendanceException.id == exc_id).first()
    if not exc:
        raise NotFoundException("AttendanceException", exc_id)
    return exc


# ── HOLIDAYS ─────────────────────────────────────────────────────────────────

def get_holidays(db: Session) -> list[Holiday]:
    return db.query(Holiday).order_by(Holiday.date).all()


def create_holiday(db: Session, data: HolidayCreate, created_by: int = None) -> Holiday:
    existing = db.query(Holiday).filter(Holiday.date == data.date, Holiday.name.ilike(data.name)).first()
    if existing:
        raise AlreadyExistsException("Holiday", "date")
    holiday = Holiday(**data.model_dump(), created_by=created_by)
    db.add(holiday)
    db.commit()
    db.refresh(holiday)
    return holiday


def get_holiday_by_id(db: Session, holiday_id: int) -> Holiday:
    holiday = db.query(Holiday).filter(Holiday.id == holiday_id).first()
    if not holiday:
        raise NotFoundException("Holiday", holiday_id)
    return holiday


def update_holiday(db: Session, holiday_id: int, data: HolidayUpdate, updated_by: int = None) -> Holiday:
    holiday = get_holiday_by_id(db, holiday_id)
    update_data = sanitize_dict(data.model_dump(exclude_unset=True))
    for field, value in update_data.items():
        setattr(holiday, field, value)
    db.commit()
    db.refresh(holiday)
    return holiday


def delete_holiday(db: Session, holiday_id: int) -> None:
    holiday = get_holiday_by_id(db, holiday_id)
    db.delete(holiday)
    db.commit()


def import_holidays(db: Session, holidays: list[dict], created_by: int = None) -> dict:
    imported = 0
    for h in holidays:
        safe = sanitize_dict(h)
        name = safe.get("name")
        h_date = safe.get("date")
        if not name or not h_date:
            continue
        existing = db.query(Holiday).filter(Holiday.date == h_date, Holiday.name.ilike(name)).first()
        if existing:
            continue
        h_type = safe.get("type", "public")
        is_recurring = safe.get("is_recurring", False)
        description = safe.get("description")
        holiday = Holiday(
            name=name,
            date=h_date,
            type=h_type,
            is_recurring=is_recurring,
            description=description,
            created_by=created_by,
        )
        db.add(holiday)
        imported += 1
    db.commit()
    return {"imported": imported}


# ── WEEKEND CONFIG ───────────────────────────────────────────────────────────

def get_weekend_configs(db: Session) -> list[WeekendConfig]:
    return db.query(WeekendConfig).order_by(WeekendConfig.day_of_week).all()


def create_weekend_config(db: Session, data: WeekendConfigCreate, created_by: int = None) -> WeekendConfig:
    existing = db.query(WeekendConfig).filter(WeekendConfig.day_of_week == data.day_of_week).first()
    if existing:
        raise AlreadyExistsException("WeekendConfig", "day_of_week")
    config = WeekendConfig(**data.model_dump(), created_by=created_by)
    db.add(config)
    db.commit()
    db.refresh(config)
    return config


def update_weekend_config(db: Session, config_id: int, data: WeekendConfigCreate, updated_by: int = None) -> WeekendConfig:
    config = db.query(WeekendConfig).filter(WeekendConfig.id == config_id).first()
    if not config:
        raise NotFoundException("WeekendConfig", config_id)
    update_data = sanitize_dict(data.model_dump(exclude_unset=True))
    for field, value in update_data.items():
        setattr(config, field, value)
    db.commit()
    db.refresh(config)
    return config


def delete_weekend_config(db: Session, config_id: int) -> None:
    config = db.query(WeekendConfig).filter(WeekendConfig.id == config_id).first()
    if not config:
        raise NotFoundException("WeekendConfig", config_id)
    db.delete(config)
    db.commit()


# ── AUDIT LOGS ───────────────────────────────────────────────────────────────

def _log_audit(db: Session, action: str, entity_type: str, entity_id: Optional[int] = None,
               employee_id: Optional[int] = None, changes: Optional[dict] = None,
               performed_by: Optional[int] = None, ip_address: Optional[str] = None) -> None:
    log = AttendanceAuditLog(
        action=action,
        entity_type=entity_type,
        entity_id=entity_id,
        employee_id=employee_id,
        changes=json.dumps(changes, default=str) if changes else None,
        performed_by=performed_by,
        ip_address=ip_address,
    )
    db.add(log)
    db.flush()


def get_audit_logs(
    db: Session, page: int = 1, per_page: int = 20,
    action: Optional[str] = None,
    entity_type: Optional[str] = None,
    employee_id: Optional[int] = None,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
) -> dict:
    per_page = min(per_page, 100)
    query = db.query(AttendanceAuditLog)
    if action:
        query = query.filter(AttendanceAuditLog.action == action)
    if entity_type:
        query = query.filter(AttendanceAuditLog.entity_type == entity_type)
    if employee_id:
        query = query.filter(AttendanceAuditLog.employee_id == employee_id)
    if date_from:
        query = query.filter(AttendanceAuditLog.created_at >= datetime.combine(date_from, datetime.min.time()))
    if date_to:
        query = query.filter(AttendanceAuditLog.created_at <= datetime.combine(date_to, datetime.max.time()))
    total = query.count()
    items = query.order_by(AttendanceAuditLog.created_at.desc()).offset((page - 1) * per_page).limit(per_page).all()
    result = []
    for r in items:
        result.append({
            "id": r.id,
            "employee_id": r.employee_id,
            "action": r.action,
            "entity_type": r.entity_type,
            "entity_id": r.entity_id,
            "changes": r.changes,
            "performed_by": r.performed_by,
            "ip_address": r.ip_address,
            "created_at": r.created_at,
            "employee_name": r.employee.full_name if r.employee else None,
            "performer_name": r.performer.full_name if r.performer else None,
        })
    return {"total": total, "page": page, "per_page": per_page, "items": result}


# ── EXPORTS ──────────────────────────────────────────────────────────────────

def export_attendance_csv(
    db: Session, search=None, status=None, department=None,
    date_from=None, date_to=None, employee_id=None,
    sort_by="date", sort_order="desc",
) -> str:
    query = _get_records_query(db, search, status, department, date_from, date_to, employee_id)
    sort_col = SORTABLE_FIELDS_RECORDS.get(sort_by, AttendanceRecord.date)
    sort_fn = desc if sort_order == "desc" else asc
    records = query.order_by(sort_fn(sort_col)).all()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["ID", "Employee ID", "Employee Name", "Date", "Status",
                     "Check In", "Check Out", "Notes", "Created At"])
    for r in records:
        writer.writerow([
            r.id, r.employee_id,
            r.employee.full_name if r.employee else "",
            r.date, r.status.value if r.status else "",
            r.check_in, r.check_out, r.notes, r.created_at,
        ])
    return output.getvalue()


def export_attendance_excel(
    db: Session, search=None, status=None, department=None,
    date_from=None, date_to=None, employee_id=None,
    sort_by="date", sort_order="desc",
) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    query = _get_records_query(db, search, status, department, date_from, date_to, employee_id)
    sort_col = SORTABLE_FIELDS_RECORDS.get(sort_by, AttendanceRecord.date)
    sort_fn = desc if sort_order == "desc" else asc
    records = query.order_by(sort_fn(sort_col)).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"

    headers = ["ID", "Employee ID", "Employee Name", "Date", "Status",
               "Check In", "Check Out", "Notes", "Created At"]
    bold = Font(bold=True)
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header).font = bold

    for row_idx, r in enumerate(records, 2):
        ws.cell(row=row_idx, column=1, value=r.id)
        ws.cell(row=row_idx, column=2, value=r.employee_id)
        ws.cell(row=row_idx, column=3, value=r.employee.full_name if r.employee else "")
        ws.cell(row=row_idx, column=4, value=str(r.date) if r.date else "")
        ws.cell(row=row_idx, column=5, value=r.status.value if r.status else "")
        ws.cell(row=row_idx, column=6, value=str(r.check_in) if r.check_in else "")
        ws.cell(row=row_idx, column=7, value=str(r.check_out) if r.check_out else "")
        ws.cell(row=row_idx, column=8, value=r.notes)
        ws.cell(row=row_idx, column=9, value=str(r.created_at) if r.created_at else "")

    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 3, 40)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def export_report_csv(db: Session, report_type: str, date_from: Optional[date] = None, date_to: Optional[date] = None) -> str:
    data = _get_report_data(db, report_type, date_from, date_to)
    output = io.StringIO()
    writer = csv.writer(output)
    if data:
        writer.writerow(data[0].keys())
        for row in data:
            writer.writerow(row.values())
    return output.getvalue()


def export_report_excel(db: Session, report_type: str, date_from: Optional[date] = None, date_to: Optional[date] = None) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    data = _get_report_data(db, report_type, date_from, date_to)
    wb = Workbook()
    ws = wb.active
    ws.title = report_type.replace("_", " ").title()

    if data:
        headers = list(data[0].keys())
        bold = Font(bold=True)
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=header).font = bold
        for row_idx, row in enumerate(data, 2):
            for col_idx, key in enumerate(headers, 1):
                ws.cell(row=row_idx, column=col_idx, value=row.get(key, ""))

    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 3, 40)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def _get_report_data(db: Session, report_type: str, date_from: Optional[date] = None, date_to: Optional[date] = None) -> list[dict]:
    query = db.query(AttendanceRecord)
    if date_from:
        query = query.filter(AttendanceRecord.date >= date_from)
    if date_to:
        query = query.filter(AttendanceRecord.date <= date_to)

    records = query.order_by(AttendanceRecord.date).all()

    if report_type == "daily":
        grouped = {}
        for r in records:
            d = str(r.date)
            if d not in grouped:
                grouped[d] = {"date": d, "present": 0, "absent": 0, "on_leave": 0, "remote": 0, "total": 0}
            grouped[d]["total"] += 1
            if r.status == AttendanceStatus.PRESENT:
                grouped[d]["present"] += 1
            elif r.status == AttendanceStatus.ABSENT:
                grouped[d]["absent"] += 1
            elif r.status == AttendanceStatus.ON_LEAVE:
                grouped[d]["on_leave"] += 1
            elif r.status == AttendanceStatus.REMOTE:
                grouped[d]["remote"] += 1
        return list(grouped.values())

    elif report_type == "monthly":
        grouped = {}
        for r in records:
            key = f"{r.date.year}-{r.date.month:02d}"
            if key not in grouped:
                grouped[key] = {"month": key, "present": 0, "absent": 0, "on_leave": 0, "remote": 0, "total": 0}
            grouped[key]["total"] += 1
            if r.status == AttendanceStatus.PRESENT:
                grouped[key]["present"] += 1
            elif r.status == AttendanceStatus.ABSENT:
                grouped[key]["absent"] += 1
            elif r.status == AttendanceStatus.ON_LEAVE:
                grouped[key]["on_leave"] += 1
            elif r.status == AttendanceStatus.REMOTE:
                grouped[key]["remote"] += 1
        return list(grouped.values())

    elif report_type == "department":
        grouped = {}
        for r in records:
            dept_name = r.employee.department.name if r.employee and r.employee.department else "Unknown"
            if dept_name not in grouped:
                grouped[dept_name] = {"department": dept_name, "present": 0, "absent": 0, "total": 0}
            grouped[dept_name]["total"] += 1
            if r.status == AttendanceStatus.PRESENT:
                grouped[dept_name]["present"] += 1
            elif r.status == AttendanceStatus.ABSENT:
                grouped[dept_name]["absent"] += 1
        return list(grouped.values())

    elif report_type == "shift":
        grouped = {}
        for r in records:
            roster = db.query(ShiftRoster).filter(
                ShiftRoster.employee_id == r.employee_id,
                ShiftRoster.date == r.date,
            ).first()
            shift_name = roster.shift.name if roster and roster.shift else "No Shift"
            if shift_name not in grouped:
                grouped[shift_name] = {"shift": shift_name, "present": 0, "absent": 0, "total": 0}
            grouped[shift_name]["total"] += 1
            if r.status == AttendanceStatus.PRESENT:
                grouped[shift_name]["present"] += 1
            elif r.status == AttendanceStatus.ABSENT:
                grouped[shift_name]["absent"] += 1
        return list(grouped.values())

    elif report_type == "late_arrivals":
        result = []
        for r in records:
            if r.status == AttendanceStatus.PRESENT and r.check_in:
                result.append({
                    "date": str(r.date),
                    "employee_id": r.employee_id,
                    "employee_name": r.employee.full_name if r.employee else "",
                    "check_in": str(r.check_in),
                })
        return result

    elif report_type == "overtime":
        ot_query = db.query(OvertimeRequest)
        if date_from:
            ot_query = ot_query.filter(OvertimeRequest.date >= date_from)
        if date_to:
            ot_query = ot_query.filter(OvertimeRequest.date <= date_to)
        ot_records = ot_query.order_by(OvertimeRequest.date).all()
        result = []
        for r in ot_records:
            result.append({
                "date": str(r.date),
                "employee_id": r.employee_id,
                "employee_name": r.employee.full_name if r.employee else "",
                "hours_requested": float(r.hours_requested),
                "hours_approved": float(r.hours_approved) if r.hours_approved else None,
                "status": r.status.value if r.status else "",
            })
        return result

    elif report_type == "absentee":
        result = []
        for r in records:
            if r.status == AttendanceStatus.ABSENT:
                result.append({
                    "date": str(r.date),
                    "employee_id": r.employee_id,
                    "employee_name": r.employee.full_name if r.employee else "",
                })
        return result

    elif report_type == "compliance":
        total_emp = db.query(func.count(Employee.id)).filter(Employee.is_active == True).scalar() or 1
        total_days = len(set(str(r.date) for r in records)) or 1
        present = sum(1 for r in records if r.status == AttendanceStatus.PRESENT)
        compliance_rate = round((present / (total_emp * total_days)) * 100, 2) if (total_emp * total_days) else 0.0
        return [{
            "total_employees": total_emp,
            "total_days": total_days,
            "total_attendance_records": len(records),
            "present_count": present,
            "compliance_rate": compliance_rate,
        }]

    return []


# ── REPORTS ──────────────────────────────────────────────────────────────────

def get_daily_report(db: Session, date_filter: Optional[date] = None) -> dict:
    d = date_filter or date.today()
    records = db.query(AttendanceRecord).filter(AttendanceRecord.date == d).all()
    present = sum(1 for r in records if r.status == AttendanceStatus.PRESENT)
    absent = sum(1 for r in records if r.status == AttendanceStatus.ABSENT)
    on_leave = sum(1 for r in records if r.status == AttendanceStatus.ON_LEAVE)
    remote = sum(1 for r in records if r.status == AttendanceStatus.REMOTE)
    return {
        "date": str(d),
        "present": present,
        "absent": absent,
        "on_leave": on_leave,
        "remote": remote,
        "total": len(records),
    }


def get_monthly_report(db: Session, month: int, year: int) -> dict:
    records = db.query(AttendanceRecord).filter(
        func.extract('month', AttendanceRecord.date) == month,
        func.extract('year', AttendanceRecord.date) == year,
    ).all()
    present = sum(1 for r in records if r.status == AttendanceStatus.PRESENT)
    absent = sum(1 for r in records if r.status == AttendanceStatus.ABSENT)
    on_leave = sum(1 for r in records if r.status == AttendanceStatus.ON_LEAVE)
    remote = sum(1 for r in records if r.status == AttendanceStatus.REMOTE)
    return {
        "month": month,
        "year": year,
        "present": present,
        "absent": absent,
        "on_leave": on_leave,
        "remote": remote,
        "total": len(records),
    }


def get_department_report(db: Session, department_name: Optional[str] = None,
                           date_from: Optional[date] = None, date_to: Optional[date] = None) -> dict:
    query = db.query(AttendanceRecord).join(Employee, AttendanceRecord.employee_id == Employee.id).join(Department)
    if department_name:
        query = query.filter(Department.name == department_name)
    if date_from:
        query = query.filter(AttendanceRecord.date >= date_from)
    if date_to:
        query = query.filter(AttendanceRecord.date <= date_to)

    records = query.all()
    dept_data = {}
    for r in records:
        dept = r.employee.department.name if r.employee and r.employee.department else "Unknown"
        if dept not in dept_data:
            dept_data[dept] = {"present": 0, "absent": 0, "total": 0}
        dept_data[dept]["total"] += 1
        if r.status == AttendanceStatus.PRESENT:
            dept_data[dept]["present"] += 1
        elif r.status == AttendanceStatus.ABSENT:
            dept_data[dept]["absent"] += 1

    breakdown = [
        {"department": d, "present": v["present"], "absent": v["absent"], "total": v["total"]}
        for d, v in dept_data.items()
    ]
    return {"department_breakdown": breakdown, "total_departments": len(breakdown)}


def get_shift_report(db: Session, shift_id: Optional[int] = None,
                      date_from: Optional[date] = None, date_to: Optional[date] = None) -> dict:
    query = db.query(ShiftRoster)
    if shift_id:
        query = query.filter(ShiftRoster.shift_id == shift_id)
    if date_from:
        query = query.filter(ShiftRoster.date >= date_from)
    if date_to:
        query = query.filter(ShiftRoster.date <= date_to)

    rosters = query.all()
    shift_data = {}
    for r in rosters:
        s_name = r.shift.name if r.shift else "Unknown"
        if s_name not in shift_data:
            shift_data[s_name] = {"total_assigned": 0, "total_present": 0}
        shift_data[s_name]["total_assigned"] += 1
        att = db.query(AttendanceRecord).filter(
            AttendanceRecord.employee_id == r.employee_id,
            AttendanceRecord.date == r.date,
            AttendanceRecord.status == AttendanceStatus.PRESENT,
        ).first()
        if att:
            shift_data[s_name]["total_present"] += 1

    breakdown = [
        {"shift": s, "total_assigned": v["total_assigned"], "total_present": v["total_present"]}
        for s, v in shift_data.items()
    ]
    return {"shift_breakdown": breakdown, "total_shifts": len(breakdown)}


def get_late_arrival_report(db: Session, date_from: Optional[date] = None, date_to: Optional[date] = None) -> dict:
    query = db.query(AttendanceRecord).filter(
        AttendanceRecord.status == AttendanceStatus.PRESENT,
        AttendanceRecord.check_in.isnot(None),
    )
    if date_from:
        query = query.filter(AttendanceRecord.date >= date_from)
    if date_to:
        query = query.filter(AttendanceRecord.date <= date_to)

    records = query.order_by(AttendanceRecord.date.desc()).all()
    result = []
    for r in records:
        result.append({
            "id": r.id,
            "employee_id": r.employee_id,
            "employee_name": r.employee.full_name if r.employee else "",
            "date": str(r.date),
            "check_in": str(r.check_in),
        })
    return {"late_arrivals": result, "total": len(result)}


def get_overtime_report_data(db: Session, date_from: Optional[date] = None, date_to: Optional[date] = None) -> dict:
    return get_overtime_reports(db, date_from, date_to)


def get_absentee_report(db: Session, date_from: Optional[date] = None, date_to: Optional[date] = None) -> dict:
    query = db.query(AttendanceRecord).filter(AttendanceRecord.status == AttendanceStatus.ABSENT)
    if date_from:
        query = query.filter(AttendanceRecord.date >= date_from)
    if date_to:
        query = query.filter(AttendanceRecord.date <= date_to)

    records = query.order_by(AttendanceRecord.date.desc()).all()
    by_employee = {}
    for r in records:
        eid = r.employee_id
        ename = r.employee.full_name if r.employee else "Unknown"
        if eid not in by_employee:
            by_employee[eid] = {"employee_id": eid, "employee_name": ename, "absent_days": 0}
        by_employee[eid]["absent_days"] += 1

    return {
        "total_absent_occurrences": len(records),
        "employee_breakdown": list(by_employee.values()),
    }


def get_compliance_report(db: Session, date_from: Optional[date] = None, date_to: Optional[date] = None) -> dict:
    total_emp = db.query(func.count(Employee.id)).filter(Employee.is_active == True).scalar() or 1

    query = db.query(AttendanceRecord)
    if date_from:
        query = query.filter(AttendanceRecord.date >= date_from)
    if date_to:
        query = query.filter(AttendanceRecord.date <= date_to)
    records = query.all()

    total_days = len(set(str(r.date) for r in records)) or 1
    present = sum(1 for r in records if r.status == AttendanceStatus.PRESENT)
    compliance_rate = round((present / (total_emp * total_days)) * 100, 2) if (total_emp * total_days) else 0.0

    return {
        "total_employees": total_emp,
        "total_days": total_days,
        "total_records": len(records),
        "present_count": present,
        "compliance_rate": compliance_rate,
    }


# ── ANALYTICS ────────────────────────────────────────────────────────────────

def get_attendance_trends(db: Session, months: int = 6) -> dict:
    today = date.today()
    start_date = today - timedelta(days=30 * months)
    records = db.query(AttendanceRecord).filter(
        AttendanceRecord.date >= start_date,
        AttendanceRecord.date <= today,
    ).order_by(AttendanceRecord.date).all()

    monthly = {}
    for r in records:
        key = f"{r.date.year}-{r.date.month:02d}"
        if key not in monthly:
            monthly[key] = {"month": key, "present": 0, "absent": 0, "on_leave": 0, "remote": 0, "total": 0}
        monthly[key]["total"] += 1
        if r.status == AttendanceStatus.PRESENT:
            monthly[key]["present"] += 1
        elif r.status == AttendanceStatus.ABSENT:
            monthly[key]["absent"] += 1
        elif r.status == AttendanceStatus.ON_LEAVE:
            monthly[key]["on_leave"] += 1
        elif r.status == AttendanceStatus.REMOTE:
            monthly[key]["remote"] += 1

    return {
        "months": list(monthly.values()),
        "total_months": len(monthly),
    }


def get_department_analysis(db: Session, date_from: Optional[date] = None, date_to: Optional[date] = None) -> dict:
    query = db.query(AttendanceRecord).join(Employee, AttendanceRecord.employee_id == Employee.id).join(Department)
    if date_from:
        query = query.filter(AttendanceRecord.date >= date_from)
    if date_to:
        query = query.filter(AttendanceRecord.date <= date_to)

    records = query.all()
    dept_data = {}
    for r in records:
        dept = r.employee.department.name if r.employee and r.employee.department else "Unknown"
        if dept not in dept_data:
            dept_data[dept] = {"present": 0, "absent": 0, "total": 0, "total_hours": 0.0}
        dept_data[dept]["total"] += 1
        if r.status == AttendanceStatus.PRESENT:
            dept_data[dept]["present"] += 1
        elif r.status == AttendanceStatus.ABSENT:
            dept_data[dept]["absent"] += 1
        if r.check_in and r.check_out:
            dept_data[dept]["total_hours"] += (r.check_out - r.check_in).total_seconds() / 3600

    breakdown = [
        {
            "department": d,
            "present": v["present"],
            "absent": v["absent"],
            "total_records": v["total"],
            "attendance_rate": round((v["present"] / max(v["total"], 1)) * 100, 2),
            "total_hours": round(v["total_hours"], 2),
        }
        for d, v in dept_data.items()
    ]
    return {"department_breakdown": breakdown, "total_departments": len(breakdown)}


def get_overtime_analytics(db: Session, date_from: Optional[date] = None, date_to: Optional[date] = None) -> dict:
    query = db.query(OvertimeRequest).filter(OvertimeRequest.status == OvertimeStatus.APPROVED)
    if date_from:
        query = query.filter(OvertimeRequest.date >= date_from)
    if date_to:
        query = query.filter(OvertimeRequest.date <= date_to)

    records = query.all()
    total_hours = sum(float(r.hours_approved or r.hours_requested) for r in records)
    by_dept = {}
    for r in records:
        dept = r.employee.department.name if r.employee and r.employee.department else "Unknown"
        if dept not in by_dept:
            by_dept[dept] = {"total_hours": 0.0, "count": 0}
        by_dept[dept]["total_hours"] += float(r.hours_approved or r.hours_requested)
        by_dept[dept]["count"] += 1

    return {
        "total_overtime_records": len(records),
        "total_overtime_hours": round(total_hours, 2),
        "department_breakdown": [
            {"department": d, "total_hours": round(v["total_hours"], 2), "count": v["count"]}
            for d, v in by_dept.items()
        ],
    }


def get_shift_efficiency(db: Session, date_from: Optional[date] = None, date_to: Optional[date] = None) -> dict:
    query = db.query(ShiftRoster)
    if date_from:
        query = query.filter(ShiftRoster.date >= date_from)
    if date_to:
        query = query.filter(ShiftRoster.date <= date_to)

    rosters = query.all()
    shift_data = {}
    for r in rosters:
        s_name = r.shift.name if r.shift else "Unknown"
        if s_name not in shift_data:
            shift_data[s_name] = {"total_assigned": 0, "total_present": 0, "total_hours": 0.0}
        shift_data[s_name]["total_assigned"] += 1

        att = db.query(AttendanceRecord).filter(
            AttendanceRecord.employee_id == r.employee_id,
            AttendanceRecord.date == r.date,
        ).first()
        if att and att.status == AttendanceStatus.PRESENT:
            shift_data[s_name]["total_present"] += 1
            if att.check_in and att.check_out:
                shift_data[s_name]["total_hours"] += (att.check_out - att.check_in).total_seconds() / 3600

    efficiency = [
        {
            "shift": s,
            "total_assigned": v["total_assigned"],
            "total_present": v["total_present"],
            "attendance_rate": round((v["total_present"] / max(v["total_assigned"], 1)) * 100, 2),
            "total_hours": round(v["total_hours"], 2),
        }
        for s, v in shift_data.items()
    ]
    return {"shift_efficiency": efficiency, "total_shifts": len(efficiency)}
